import random
import openpyxl
import json
import os
from get_ascii_art import get_ascii_art  # Import ASCII art function

# Dungeon Generator
# Copyright (c) 2024 Sisco1138
# Licensed under the MIT License
# See the LICENSE file or https://opensource.org/licenses/MIT for full details.

"""
Dungeon Generator is a Python tool for creating randomized dungeon layouts,
ideal for RPG campaigns, game development, and procedural storytelling.

This program allows users to generate rooms, hallways, traps, and treasures
and save or load layouts for ongoing use.
"""

# Lists for dungeon components
hallway = [
    'Straight_Hallway_Horizontal', 'Straight_Hallway_Vertical', 'L_Hallway_East_South', "L_Hallway_North_West", 
    'L_Hallway_North_East', 'L_Hallway_West_South', 'L_Hallway_South_East', 'L_Hallway_West_North','L_Hallway_South_West',
    'L_Hallway_Easr_North', 'T_Junction_Upward' 'T_Junction_Downward', 'T_Junction_Left', 'T_Junction_Right',
    'T_Junction_Right', 'Crossroad', 'Dead_End_Hallway_North', 'Dead_End_Hallway_South', 'Dead_End_Hallway_East',
    'Dead_End_Hallway_West'
]
room = [
    'Shrine/Altar Room', 'Single_Door_Room_North', 'Single_Door_Room_South', 'Single_Door_Room_East', 
    'Single_Door_Room_West', 'Two_Door_Room_North_and_South', 'Two_Door_Room_East_and_West', 'Two_Door_Room_North_and_West', 
    'Two_Door_Room_North_and_East', 'Two_Door_Room_South_and_East', 'Two_Door_Room_South_and_West','Shrine_Room_North_and_West',
    'Shrine_Room_North_and_East', 'shrine_Room_South_and_East', 'Shrine_Room_South_and_West', 'Four-Door Room_North_South_East_and_West',
    'Boss_Room_Single_Door_Room_North', 'Boss_Room_Single_Door_Room_South', 'Boss_Room_Single_Door_Room_East', 'Boss_Room_Single_Door_Room_West',
    'Boss_Room_Two_Door_Room_North_and_South', 'Boss_Room_Two_Door_Room_East_and_West', 'Boss_Room_Two_Door_Room_North_and_West',
    'Boss_Room_Two_Door_Room_North_and_East', 'Boss_Room_Two_Door_Room_South_and_East', 'Boss_Room_Two_Door_Room_South_and_West',
    'Boss_Room_Corner_Room_North_and_West', 'Boss_Room_Corner_Room_North_and_East', 'Boss_Room_Corner_Room_South_and_East',
    'Boss_Room_Corner_Room_South_and_West', 'Boss_Room_Four-Door Room_North_South_East_and_West', 'Treasure_Room_Single_Door_Room_North',
    'Treasure_Room_Single_Door_Room_South', 'Treasure_Room_Single_Door_Room_East', 'Treasure_Room_Single_Door_Room_West', 'Treasure_Room_Two_Door_Room_North_and_South',
    'Treasure_Room_Two_Door_Room_East_and_West', 'Treasure_Room_Two_Door_Room_North_and_West', 'Treasure_Room_Two_Door_Room_North_and_East',
    'Treasure_Room_Two_Door_Room_South_and_East', 'Treasure_Room_Two_Door_Room_South_and_West', 'Treasure_Room_Corner_Room_North_and_West',
    'Treasure_Room_Corner_Room_North_and_East', 'Treasure_Room_Corner_Room_South_and_East', 'Treasure_Room_Corner_Room_South_and_West'
]
trap = [
    'Illusion/Disorientation Traps', 'Poisonous Gas Traps', 
    'Swinging Blades/Pendulums', 'Arrow/Dart Traps', 
    'Pitfalls', 'Spike Traps'
]
treasures = [
    'Pile of Gold Coins', 'Rare Gemstone', 'Ancient Sword', 
    'Potion of Healing', 'Enchanted Ring', 'Magic Scroll', 
    'Mysterious Artifact'
]
descriptions = [
    "flickering torchlight", "damp and musty", "echoes of distant footsteps", 
    "crumbling walls", "covered in cobwebs", "filled with a mysterious fog"
]

# Initialize the dungeon layout with coordinates
dungeon_layout = {}

# Function to load a dungeon layout from JSON
def load_dungeon(name):
    filename = f"{name}.json"
    if os.path.exists(filename):
        with open(filename, 'r') as f:
            dungeon_data = json.load(f)
            # Convert string keys back to tuples for use in the program
            dungeon_data["layout"] = {tuple(map(int, key.split(','))): value for key, value in dungeon_data["layout"].items()}
            print(f"Dungeon '{name}' loaded successfully!")
            return dungeon_data
    else:
        print("Dungeon not found. Starting a new one.")
        return {"name": name, "layout": {}}

# Function to save a dungeon layout to JSON
def save_dungeon(dungeon_data):
    layout_with_string_keys = {f"{x},{y}": desc for (x, y), desc in dungeon_layout.items()}
    dungeon_data["layout"] = layout_with_string_keys
    
    filename = f"{dungeon_data['name']}.json"
    with open(filename, 'w') as f:
        json.dump(dungeon_data, f, indent=4)
    print(f"Dungeon '{dungeon_data['name']}' saved successfully!")

# Function to add room with coordinates
# Function to add room with coordinates
def add_to_dungeon(component, description):
    while True:
        try:
            x, y = map(int, input("Enter coordinates as x,y (negative values allowed): ").split(","))
            dungeon_layout[(x, y)] = f"{component} with {description}"
            # Display the ASCII art after adding the component
            ascii_key = f"{component.replace(' ', '_')}_{description.replace(' ', '_')}"
            print(get_ascii_art(ascii_key))  # Corrected to call get_ascii_art, not ASCII_ART directly
            break
        except ValueError:
            print("Invalid input. Please enter coordinates in the format x,y (e.g., 0,1 or -1,2).")

# Function to display ASCII map
def display_ascii_map():
    max_x = max((x for x, y in dungeon_layout), default=0)
    max_y = max((y for x, y in dungeon_layout), default=0)
    
    for y in range(max_y + 1):
        row = ""
        for x in range(max_x + 1):
            # Retrieve room or hallway at coordinates
            component = dungeon_layout.get((x, y), None)
            if component:
                # Extract the key without description for ASCII art
                component_key = component.split(" with ")[0]
                ascii_art = get_ascii_art(component_key)
                row += ascii_art if ascii_art != "ASCII art not found" else "[ ]"  # Placeholder if art is missing
            else:
                row += "[ ]"  # Empty space for coordinates without components
        print(row)

# Function to export layout to Excel with full ASCII art in a single cell
def export_to_excel(filename="dungeon_map.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    for (x, y), component in dungeon_layout.items():
        # Get the key without description to retrieve ASCII art
        component_key = component.split(" with ")[0]
        ascii_art = get_ascii_art(component_key)
        
        # If ASCII art is found, join all lines into a single string with newlines
        if ascii_art != "ASCII art not found":
            ascii_art_single_cell = "\n".join(ascii_art.strip().splitlines())
            cell = ws.cell(row=(y + 1), column=(x + 1))
            cell.value = ascii_art_single_cell
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)  # Enable text wrapping
        else:
            ws.cell(row=(y + 1), column=(x + 1), value="[ ]")  # Placeholder if ASCII art is missing
    
    # Save the workbook
    wb.save(filename)
    print(f"Dungeon layout with ASCII art in single cells exported to {filename}")
    
    # Calculate the max dimensions for the layout
    max_x = max((x for x, y in dungeon_layout), default=0)
    max_y = max((y for x, y in dungeon_layout), default=0)
    
    for (x, y), component in dungeon_layout.items():
        # Get the key without description to retrieve ASCII art
        component_key = component.split(" with ")[0]
        ascii_art = get_ascii_art(component_key)
        
        # If ASCII art is found, write each line to consecutive rows in the cell
        if ascii_art != "ASCII art not found":
            # Split ASCII art by lines
            ascii_lines = ascii_art.strip().splitlines()
            for i, line in enumerate(ascii_lines):
                # Write each line of ASCII art in consecutive cells starting from (y+1, x+1)
                ws.cell(row=(y * len(ascii_lines)) + 1 + i, column=(x + 1), value=line)
        else:
            ws.cell(row=(y + 1), column=(x + 1), value="[ ]")  # Placeholder if ASCII art is missing
    
    # Save the workbook
    wb.save(filename)
    print(f"Dungeon layout with ASCII art exported to {filename}")

# Initialize the dungeon
dungeon_name = input("Enter a name for your dungeon: ").strip()
dungeon_data = load_dungeon(dungeon_name)

print("Welcome to the Dungeon Generator!")

while True:
    print("\nWhat would you like to add to your dungeon? (hallway, room, trap, treasure, def for definitions, layout to review, map to show ASCII map, excel to export, save to save, or quit to exit)")
    choice = input(": ").strip().lower()
    
    if choice == "hallway":
        while True:
            hallway_choice = random.choice(hallway)
            hallway_description = random.choice(descriptions)
            print("The next part of your dungeon is:", f"{hallway_choice} with {hallway_description}")
            
            reroll = input("Keep this hallway? Type 'y' to keep or 'n' to re-roll: ").strip().lower()
            if reroll == 'y':
                add_to_dungeon(hallway_choice, hallway_description)
                break
            elif reroll != 'n':
                print("Invalid input. Please enter 'y' to keep or 'n' to re-roll.")

    elif choice == "room":
        while True:
            room_choice = random.choice(room)
            room_description = random.choice(descriptions)
            print("The next part of your dungeon is:", f"{room_choice} with {room_description}")
            
            reroll = input("Keep this room? Type 'y' to keep or 'n' to re-roll: ").strip().lower()
            if reroll == 'y':
                add_to_dungeon(room_choice, room_description)
                break
            elif reroll != 'n':
                print("Invalid input. Please enter 'y' to keep or 'n' to re-roll.")

    elif choice == "trap":
        while True:
            trap_choice = random.choice(trap)
            print("The next part of your dungeon is a trap:", trap_choice)
            
            reroll = input("Keep this trap? Type 'y' to keep or 'n' to re-roll: ").strip().lower()
            if reroll == 'y':
                add_to_dungeon("Trap", trap_choice)
                break
            elif reroll != 'n':
                print("Invalid input. Please enter 'y' to keep or 'n' to re-roll.")

    elif choice == "treasure":
        while True:
            treasure_choice = random.choice(treasures)
            print("You've found a treasure:", treasure_choice)
            
            reroll = input("Keep this treasure? Type 'y' to keep or 'n' to re-roll: ").strip().lower()
            if reroll == 'y':
                add_to_dungeon("Treasure", treasure_choice)
                break
            elif reroll != 'n':
                print("Invalid input. Please enter 'y' to keep or 'n' to re-roll.")
    
    elif choice == "layout":
        print("\nDungeon Layout So Far:")
        for coordinates, description in dungeon_layout.items():
            print(f"At {coordinates}: {description}")
    
    elif choice == "map":
        print("\nASCII Map of the Dungeon:")
        display_ascii_map()
    
    elif choice == "excel":
        export_to_excel()
    
    elif choice == 'def':
        print("""
        ... [Full description of rooms, hallways, and traps here]
        """)
    
    elif choice == "save":
        dungeon_data["layout"] = dungeon_layout
        save_dungeon(dungeon_data)
    
    elif choice == "quit":
        print("Thank you for using the Dungeon Generator!")
        break
    
    else: 
        print("That is not an option. Please select one of the options: hallway, room, trap, treasure, def, layout, map, excel, save, or quit.")
